import os
import shutil
import openpyxl
import datetime
import asyncio


def create_xls(user_id):
    """
    Функция для создания Excel файла для пользователя
    :param user_id: ID пользователя, для которого создается Excel-файл

    :type user_id: int

    :return: Функция создает папку для пользователя (если она еще не существует), копирует шаблон Excel-файла
    в эту папку и называет его по ID пользователя.

    Логика работы:
    1. Определяется папка назначения на основе ID пользователя. Путь файла назначения — это комбинация папки пользователя
    и его ID.
    2. Если папка для пользователя не существует, она создается с помощью `os.makedirs()`.
    3. Шаблонный Excel-файл копируется в папку пользователя под его ID с помощью функции `shutil.copy2()`.
    4. После успешного копирования выводится сообщение о том, что папка создана и файл скопирован.

    Используемые методы:
    - `os.path.exists()`: Проверяет наличие папки пользователя.
    - `os.makedirs()`: Создает новую папку для пользователя, если она не существует.
    - `shutil.copy2()`: Копирует исходный файл в папку назначения, сохраняя метаданные файла.

    Примечание:
    - Шаблонный файл Excel называется 'Простой бюджет на месяц1.xlsx' и должен находиться в той же директории, откуда запускается скрипт.
    - Путь для папки пользователя создается в формате "user_files/{user_id}/", а сам Excel-файл сохраняется под именем "{user_id}.xlsx".
    """

    source = 'Простой бюджет на месяц1.xlsx'
    destination_folder = f"user_files/{user_id}/"
    destination = f"{destination_folder}/{user_id}.xlsx"

    # Создание папки, если она не существует
    if not os.path.exists(destination_folder):
        os.makedirs(destination_folder)

    shutil.copy2(source, destination)
    print('Папка создана и файл скопирован')


async def get_cell_value(cell_address, user_id, sheet_name='2024'):
    """
    Асинхронная функция для получения значения из ячейки Excel-файла
    :param cell_address: Адрес ячейки, из которой нужно получить значение (например, 'A1')
    :param user_id: ID пользователя, для которого ищется файл Excel
    :param sheet_name: Имя листа Excel, с которого нужно получить значение (по умолчанию используется лист '2024')

    :type cell_address: str
    :type user_id: int
    :type sheet_name: str

    :return: Возвращает значение ячейки и объект workbook, представляющий Excel-файл.

    Логика работы:
    1. Определяется путь к файлу Excel на основе ID пользователя. Файл должен находиться в папке "user_files/{user_id}/"
    и называться "{user_id}.xlsx".
    2. Асинхронно открывается файл Excel с помощью `openpyxl.load_workbook()` через вызов `asyncio.to_thread()`, чтобы не блокировать основной поток.
    3. Из загруженной книги выбирается лист по имени, указанному в параметре `sheet_name` (по умолчанию — '2024').
    4. Из выбранного листа извлекается значение указанной ячейки с помощью `sheet[cell_address].value`.
    5. Возвращается кортеж, содержащий значение ячейки и объект workbook.

    Используемые методы:
    - `asyncio.to_thread()`: Выполняет блокирующие операции (например, чтение файла) в отдельном потоке.
    - `openpyxl.load_workbook()`: Загружает книгу Excel для работы с ней.
    - `sheet[cell_address].value`: Извлекает значение указанной ячейки из листа.

    Примечание:
    - Если лист с именем, указанным в `sheet_name`, не существует, будет выброшено исключение.
    - Убедитесь, что файл Excel существует в указанной папке и имеет правильный формат.
    """

    file_path = f'user_files/{user_id}/{user_id}.xlsx'
    # Открываем файл Excel асинхронно
    workbook = await asyncio.to_thread(openpyxl.load_workbook, file_path)

    # Выбираем лист по имени
    sheet = workbook[sheet_name]

    # Получаем значение ячейки
    cell_value = sheet[cell_address].value

    return cell_value, workbook


async def add_value_to_cell(cell_address, value_to_add, user_id, sheet_name='2024'):
    file_path = f'user_files/{user_id}/{user_id}.xlsx'
    """
    Функция для добавления значения к ячейке и сохранения изменений в файл.

    :param user_id: Имя файла которое схоже с id пользователя
    :param sheet_name: Название листа
    :param cell_address: Адрес ячейки в формате A1, B2 и т.д.
    :param value_to_add: Значение, которое нужно прибавить к существующему значению ячейки
    """
    # Получаем текущее значение ячейки и объект workbook
    cell_value, workbook = await get_cell_value(cell_address, user_id, sheet_name)

    # Если значение в ячейке пустое, считать его за 0
    if cell_value is None:
        cell_value = 0

    # Добавляем значение
    new_value = cell_value + value_to_add

    # Обновляем значение в ячейке
    sheet = workbook[sheet_name]
    sheet[cell_address].value = new_value

    # Сохраняем изменения в файл асинхронно
    await asyncio.to_thread(workbook.save, file_path)


async def data_validator(user_id, button_type, amount):
    month = datetime.datetime.now().month
    buttons = {
        "Зп на руки": '3', "Зп на карточку": '4', "Шабашки": '5', "Другие": '6',
        "Жилье": '13', "Коммуналка": '14',
        "Еда": '15', "Проезд": '16',
        "Интернет": '17', "Сотовая связь": '18',
        "Одежда": '19', "Медикаменты": '20',
        "Процент кредита": '21', "Хоз расходы": '22',
        "Техника": '23', "Парикмахерская": '24',
        "Развлечения": '25', "Обучение": '26',
        "Подарки": '27', "Прочие": '28'
    }
    month_id = {
        1: 'H', 2: 'I', 3: 'J', 4: 'K',
        5: 'L', 6: 'M', 7: 'N', 8: 'C',
        9: 'D', 10: 'E', 11: 'F', 12: 'G'
    }
    await add_value_to_cell(cell_address=f'{month_id[month]}{buttons[button_type]}', value_to_add=amount,
                            user_id=user_id)
